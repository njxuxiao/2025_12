import pandas as pd
import os
import sys
from openpyxl import load_workbook
import tempfile
import time

# 尝试导入 keyboard 库
try:
    import keyboard
except ImportError:
    print("错误：缺少 'keyboard' 库。")
    print("请在你的终端或命令行中运行: pip install keyboard")
    sys.exit()

# 【方案变更】同时支持 gTTS (在线) 和 pyttsx3 (离线)
try:
    from gtts import gTTS
    from playsound import playsound
    gtts_available = True
except ImportError:
    print("警告：未安装 gTTS 或 playsound 库，将无法使用在线朗读功能。")
    print("  - pip install gTTS")
    print("  - pip install playsound==1.2.2")
    gtts_available = False

try:
    import pyttsx3
    pyttsx3_available = True
except ImportError:
    print("警告：未安装 pyttsx3 库，将无法使用离线备用朗读功能。")
    print("  - pip install pyttsx3")
    pyttsx3_available = False


def get_pyttsx3_japanese_voice_id():
    """
    【BUG修复】扫描并返回日语语音的ID，而不是返回引擎实例。
    这样可以为每次朗读创建新的引擎，避免状态问题。
    """
    if not pyttsx3_available:
        return None
    try:
        # 使用一个临时引擎来扫描可用的语音
        temp_engine = pyttsx3.init()
        voices = temp_engine.getProperty('voices')
        temp_engine.stop() # 立即停止并清理临时引擎
        del temp_engine

        japanese_voice_id = None
        for voice in voices:
            lang_str = getattr(voice, 'lang', '').lower()
            name_str = getattr(voice, 'name', '').lower()
            if 'ja' in lang_str or 'japanese' in name_str:
                japanese_voice_id = voice.id
                break
        
        if japanese_voice_id:
            print("\n成功找到【本地】日语语音ID，备用方案可用。")
            return japanese_voice_id
        else:
            print("\n警告：未在您的系统中找到日语【本地】语音包，离线备用方案可能无法朗读。")
            return None
    except Exception as e:
        print(f"\n扫描本地语音时失败: {e}")
        return None


def speak_with_gtts(text):
    """【在线】使用 gTTS 朗读，失败时返回 False。"""
    if not gtts_available or not text:
        return False
    
    temp_mp3 = None
    fp = None
    success = False

    try:
        fp = tempfile.NamedTemporaryFile(delete=False, suffix='.mp3')
        temp_mp3 = fp.name
        fp.close()
        fp = None

        for attempt in range(3):
            try:
                tts = gTTS(text=text, lang='ja')
                tts.save(temp_mp3)
                success = True
                break 
            except Exception as e:
                print(f"\n!! 在线朗读连接失败 (尝试 {attempt + 1}/3): {e}")
                if attempt < 2:
                    time.sleep(0.5)

        if success:
            playsound(temp_mp3)
        return success
    except Exception as e:
        print(f"\n!! 朗读时发生未知错误: {e}")
        return False
    finally:
        if fp is not None:
            fp.close()
        if temp_mp3 and os.path.exists(temp_mp3):
            try:
                os.remove(temp_mp3)
            except Exception as e:
                print(f"!! 清理临时文件失败: {e}")


def speak_with_pyttsx3(voice_id, text):
    """
    【离线】【BUG修复】为每次朗读创建一个全新的引擎实例，以避免状态错误。
    """
    if voice_id and text:
        try:
            # 为每个单词创建一个新的引擎实例
            engine = pyttsx3.init()
            engine.setProperty('voice', voice_id)
            engine.say(text)
            engine.runAndWait()
            engine.stop() # 确保资源被释放
        except Exception as e:
            print(f"\n!! 使用本地语音朗读时出错: {e}")


def get_display_length(s):
    """计算字符串在终端中的显示宽度（中文算2，英文算1）"""
    length = 0
    for char in str(s):
        if '\u4e00' <= char <= '\u9fff':
            length += 2
        else:
            length += 1
    return length

def display_term(word, grammar):
    """美化输出，让单词和语法在控制台中更显眼"""
    print("\n" + "╔" + "═"*50 + "╗")
    print("║" + " "*50 + "║")
    
    if word:
        line = f"  【单词】: {word}"
        padding = 50 - get_display_length(line)
        if padding < 0: padding = 0
        print("║" + line + " "*padding + "║")
        
    if grammar:
        line = f"  【文法】: {grammar}"
        padding = 50 - get_display_length(line)
        if padding < 0: padding = 0
        print("║" + line + " "*padding + "║")

    print("║" + " "*50 + "║")
    print("╚" + "═"*50 + "╝")

def display_details(meaning, remarks):
    """美化输出，用于显示含义和备注，使用虚线边框"""
    if not meaning and not remarks:
        return

    print("╭" + "┈"*50 + "╮")
    
    if meaning:
        line = f"  【含义】: {meaning}"
        padding = 50 - get_display_length(line)
        if padding < 0: padding = 0
        print("┆" + line + " "*padding + "┆")

    if remarks:
        line = f"  【备注】: {remarks}"
        padding = 50 - get_display_length(line)
        if padding < 0: padding = 0
        print("┆" + line + " "*padding + "┆")

    print("╰" + "┈"*50 + "╯")


def study_helper(file_path, sheet_to_study=None, tts_mode='auto'):
    """
    一个帮助你从Excel文件中学习日语单词和语法的脚本。
    它会逐个展示单词/语法，如果你回答“不记得”，
    程序会自动在 'Fre' 列中为该词条计数。

    Args:
        file_path (str): 你的Excel文件的路径。
        sheet_to_study (int or str, optional): 指定要学习的工作表 (索引或名称)。默认为 None (手动选择)。
        tts_mode (str): 指定语音引擎模式 ('auto', 'online', 'offline')。
    """
    
    # --- 初始化双核语音引擎 ---
    japanese_voice_id = get_pyttsx3_japanese_voice_id()
    # 仅用于 'auto' 模式下的状态跟踪
    auto_mode_current_engine = 'gTTS' if gtts_available else 'pyttsx3'

    # --- 1. 读取Excel文件并选择工作表 (Sheet) ---
    all_sheets_data = {}
    chosen_sheet = None
    try:
        if not os.path.exists(file_path):
            print(f"错误：找不到文件 '{file_path}'。")
            return

        if not file_path.endswith('.xlsx'):
            print(f"错误：此功能需要使用 .xlsx 格式的Excel文件。")
            print(f"请用Excel打开 '{os.path.basename(file_path)}' 并另存为 .xlsx 格式。")
            return
        
        xls = pd.ExcelFile(file_path, engine='openpyxl')
        sheet_names = xls.sheet_names

        if not sheet_names:
            print("错误：Excel文件中没有任何工作表。")
            return

        if sheet_to_study is not None:
            if isinstance(sheet_to_study, int):
                if 0 <= sheet_to_study < len(sheet_names):
                    chosen_sheet = sheet_names[sheet_to_study]
                else:
                    print(f"错误：指定的工作表索引 {sheet_to_study} 无效。有效索引范围是 0 到 {len(sheet_names)-1}。")
                    return
            elif isinstance(sheet_to_study, str):
                if sheet_to_study in sheet_names:
                    chosen_sheet = sheet_to_study
                else:
                    print(f"错误：找不到名为 '{sheet_to_study}' 的工作表。")
                    print(f"可用的工作表有: {sheet_names}")
                    return
            else:
                print("错误：提供的 sheet_to_study 参数类型不正确，应为整数（索引）或字符串（名称）。")
                return
            print(f"根据指定，已选择工作表: '{chosen_sheet}'")
        else:
            if len(sheet_names) == 1:
                chosen_sheet = sheet_names[0]
                print(f"自动选择唯一的工作表: '{chosen_sheet}'")
            else:
                print("发现多个工作表 (Sheet):")
                for i, name in enumerate(sheet_names):
                    print(f"  {i+1}: {name}")
                while True:
                    try:
                        choice = int(input(f"请输入你想学习的工作表编号 (1-{len(sheet_names)}): "))
                        if 1 <= choice <= len(sheet_names):
                            chosen_sheet = sheet_names[choice-1]
                            break
                        else:
                            print("编号无效，请重新输入。")
                    except ValueError:
                        print("请输入数字。")
        
        all_sheets_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        df = all_sheets_data[chosen_sheet]

    except Exception as e:
        print(f"读取或选择工作表时发生错误: {e}")
        return

    # --- 2. 准备数据 ---
    if 'Fre' not in df.columns:
        print("检测到没有'Fre'列，已自动为您创建。")
        df['Fre'] = 0
    else:
        df['Fre'] = pd.to_numeric(df['Fre'], errors='coerce').fillna(0).astype(int)

    if '单词' not in df.columns and '文法' not in df.columns:
        print(f"错误：工作表 '{chosen_sheet}' 中必须至少包含 '单词' 或 '文法' 列。")
        return
        
    has_meaning_col = '含义' in df.columns
    has_remarks_col = '备注' in df.columns
    if not has_meaning_col:
        print("提示：你的Excel中没有“含义”列，将不会显示解释。")
    if not has_remarks_col:
        print("提示：你的Excel中没有“备注”列，将不会显示备注信息。")

    df.sort_values(by='Fre', ascending=False, inplace=True)
    print("\n已按照“忘记频率”排序，忘记次数最多的将优先学习。")

    print("\n--- 日语学习助手已启动 ---")
    # 【新】根据模式打印提示
    if tts_mode == 'online':
        print("【语音模式】：仅在线")
    elif tts_mode == 'offline':
        print("【语音模式】：仅本地")
    else:
        print("【语音模式】：自动 (优先在线，失败后切换)")
        
    print("【重要】请确保你的输入法是英文模式，以便程序能正确识别按键。")
    
    # --- 3. 学习循环 ---
    is_changed = False
    last_answered_correctly_index = None
    records = df.to_dict('records')
    original_indices = df.index.tolist()

    i = 0
    while i < len(records):
        current_record = records[i]
        original_index = original_indices[i]

        word = str(current_record.get('单词', '')) if pd.notna(current_record.get('单词')) else ""
        grammar = str(current_record.get('文法', '')) if pd.notna(current_record.get('文法')) else ""
        meaning = str(current_record.get('含义', '')) if has_meaning_col and pd.notna(current_record.get('含义')) else ""
        remarks = str(current_record.get('备注', '')) if has_remarks_col and pd.notna(current_record.get('备注')) else ""

        if not word and not grammar:
            i += 1
            continue
        
        display_term(word, grammar)

        prompt = "请按键... (→: 记得 / 0: 不记得 / q: 退出"
        if last_answered_correctly_index is not None:
            prompt += " / x: 更正上一题)"
        else:
            prompt += ")"
        print(prompt)
        
        event = keyboard.read_event(suppress=True)
        while event.event_type != keyboard.KEY_DOWN:
            event = keyboard.read_event(suppress=True)
        
        key = event.name.lower()
        
        if key == 'q':
            print("正在保存进度并退出...")
            break 
        
        if key == 'x':
            if last_answered_correctly_index is not None:
                for idx, record in enumerate(records):
                    if original_indices[idx] == last_answered_correctly_index:
                        record['Fre'] += 1
                        break
                is_changed = True
                print(f"\n已对上一题进行更正！")
                last_answered_correctly_index = None
            else:
                print("\n没有可以更正的上一题。")
            continue

        last_answered_correctly_index = None
        text_to_speak = word if word else grammar

        # 【新】根据 tts_mode 调用不同的语音引擎逻辑
        def speak():
            if tts_mode == 'online':
                if not speak_with_gtts(text_to_speak):
                    print("\n!! 在线朗读失败。")
                return

            if tts_mode == 'offline':
                if japanese_voice_id:
                    speak_with_pyttsx3(japanese_voice_id, text_to_speak)
                else:
                    print("\n!! 本地语音不可用。")
                return
            
            # 默认执行 'auto' 逻辑
            nonlocal auto_mode_current_engine
            if auto_mode_current_engine == 'gTTS':
                if not speak_with_gtts(text_to_speak):
                    print("\n!! 在线朗读失败，已自动切换到【本地语音】模式。")
                    auto_mode_current_engine = 'pyttsx3'
                    speak_with_pyttsx3(japanese_voice_id, text_to_speak)
            elif auto_mode_current_engine == 'pyttsx3':
                speak_with_pyttsx3(japanese_voice_id, text_to_speak)

        if key == '0':
            current_record['Fre'] += 1
            is_changed = True
            print(f"已记录！忘记次数: {current_record['Fre']}")
            display_details(meaning, remarks)
            speak()

        elif key == 'right':
            display_details(meaning, remarks)
            speak()
            print("很好！")
            last_answered_correctly_index = original_index

        i += 1

    print("\n所有单词/语法已学习完毕！")

    # --- 5. 保存数据到Excel ---
    if not is_changed:
        print("\n内容无变化，无需保存。")
        return
        
    try:
        print("正在将学习记录更新回DataFrame...")
        new_df = pd.DataFrame(records, index=original_indices)
        all_sheets_data[chosen_sheet] = new_df

        print("正在保存文件并保留列宽...")
        book = load_workbook(file_path)
        col_widths = {}
        for sheet_name in book.sheetnames:
            col_widths[sheet_name] = {
                letter: dim.width for letter, dim in book[sheet_name].column_dimensions.items()
            }
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, sheet_data in all_sheets_data.items():
                sheet_data.to_excel(writer, sheet_name=sheet_name, index=True)
                
                if sheet_name in col_widths:
                    ws = writer.sheets[sheet_name]
                    for col_letter, width in col_widths[sheet_name].items():
                        if width:
                           ws.column_dimensions[col_letter].width = width

        print("\n学习结束！你的进度已成功保存，且列宽保持不变。")
    except PermissionError:
        print(f"\n保存文件时发生错误：权限不足。请关闭Excel文件 '{file_path}' 后重试。")
    except Exception as e:
        print(f"\n保存文件时发生未知错误: {e}")


# --- 使用说明 ---
# 1. 【重要】请确保你的Excel文件是 .xlsx 格式。
# 2. 【重要】请确保你已经安装了所需的库：
#    - pip install pandas
#    - pip install openpyxl
#    - pip install keyboard
#    - pip install gTTS              <- 在线语音库
#    - pip install playsound==1.2.2  <- 播放库
#    - pip install pyttsx3           <- 离线备用语音库
#
# 3. 【重要】在某些操作系统上，由于脚本需要监听键盘，你可能需要【以管理员权限】来运行它。
#
# 4. 将下面的 '你的日语单词.xlsx' 替换成你的Excel文件名或完整路径。
if __name__ == '__main__':
    excel_file_path = './21_7/21_7.xlsx' 
    
    study_sheet = 0

    # 'auto'
    # 'online'
    # 'offline'
    preferred_tts_engine = 'offline' 

    study_helper(excel_file_path, sheet_to_study=study_sheet, tts_mode=preferred_tts_engine)

